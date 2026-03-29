"""
data_reader.py
--------------
DataReader — reads test credentials and parameters from validdata.xlsx.
Interface Segregation: sole responsibility is Excel data extraction.

TC-83 Data Source (validdata.xlsx):
    Row 1 Headers : username | password
    Row 2 Values  : standard_user | secret_sauce

Usage:
    from utils.data_reader import DataReader
    creds = DataReader.get_valid_credentials()
    username = creds["username"]   # standard_user
    password = creds["password"]   # secret_sauce
"""

import os
from utils.logger import Logger

logger = Logger.get_logger(__name__)

# Path to the Excel data file (relative to project root)
DATA_FILE_PATH = os.path.join(os.path.dirname(__file__), "..", "validdata.xlsx")


class DataReader:
    """Utility for reading test data from validdata.xlsx."""

    @staticmethod
    def get_valid_credentials(
        file_path: str = DATA_FILE_PATH,
        sheet_name: str = "Sheet1",
        row: int = 1,
    ) -> dict:
        """
        Read valid login credentials from validdata.xlsx.

        Parameters
        ----------
        file_path : str
            Path to the Excel file (default: validdata.xlsx in project root).
        sheet_name : str
            Sheet name to read from (default: 'Sheet1').
        row : int
            Data row index (1-indexed, excludes header row; default: 1).

        Returns
        -------
        dict
            {'username': 'standard_user', 'password': 'secret_sauce'}

        Notes
        -----
        Falls back to Config.USERNAME / Config.PASSWORD if openpyxl is
        unavailable or the file is not found.
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            sheet    = workbook[sheet_name]

            # Row 1 = headers; data starts at row 2
            data_row = row + 1
            username = sheet.cell(row=data_row, column=1).value
            password = sheet.cell(row=data_row, column=2).value

            logger.info(
                f"[DataReader] Loaded credentials from '{file_path}' "
                f"sheet='{sheet_name}' row={data_row} → user='{username}'"
            )
            return {"username": username, "password": password}

        except FileNotFoundError:
            logger.warning(
                f"[DataReader] '{file_path}' not found — "
                "falling back to Config defaults"
            )
        except Exception as exc:
            logger.warning(
                f"[DataReader] Could not read '{file_path}': {exc} — "
                "falling back to Config defaults"
            )

        # Graceful fallback to Config constants
        from utils.config import Config
        return {"username": Config.USERNAME, "password": Config.PASSWORD}

    @staticmethod
    def get_all_credentials(
        file_path: str = DATA_FILE_PATH,
        sheet_name: str = "Sheet1",
    ) -> list:
        """
        Read all credential rows from validdata.xlsx.

        Returns
        -------
        list of dict
            [{'username': ..., 'password': ...}, ...]
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            sheet    = workbook[sheet_name]

            credentials = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    credentials.append({"username": row[0], "password": row[1]})

            logger.info(
                f"[DataReader] Loaded {len(credentials)} credential rows "
                f"from '{file_path}'"
            )
            return credentials

        except Exception as exc:
            logger.warning(f"[DataReader] Could not read all rows: {exc}")
            from utils.config import Config
            return [{"username": Config.USERNAME, "password": Config.PASSWORD}]
